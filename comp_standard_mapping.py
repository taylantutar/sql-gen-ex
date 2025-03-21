import io
from openpyxl import load_workbook
from itertools import chain


async def MappingDataPrepear(file):
    try:
        contents = await file.read()
        workbook = load_workbook(io.BytesIO(contents), read_only=True)
        sheet = workbook.active

        mappingList = []
        for i, row in enumerate(sheet.iter_rows()):
            if (i == 0):
                continue
            if (row[0].value == None):
                break
            mappingDict = {}
            mappingDict["SrcStandardCode"] = row[0].value.strip()
            mappingDict["SrcItemCode"] = row[1].value.strip()
            mappingDict["DestStandardCode"] = row[2].value.strip()
            mappingDict["DestItemCode"] = row[3].value.strip()
            mappingDict["BertMeanStr"] = row[4].value
            mappingDict["BertMean"] = int(round(float(row[4].value) * 100, 0))
            mappingList.append(mappingDict)

        return mappingList

    except FileNotFoundError:
        raise Exception(f"Error: {file.filename} file is not found.")
    except Exception as e:
        raise Exception(f"An error occured: {e}")
 

def CheckValidations(mappingList, sFrmList, iFrmList):
    try:
        allDbFrmList = sFrmList+ iFrmList

        srcFrmList = [row["SrcStandardCode"] for row in mappingList]
        destFrmList = [row["DestStandardCode"] for row in mappingList]
        allFrmList = set(srcFrmList + destFrmList)

        exceptedFtmList = list(set(allFrmList) - set(allDbFrmList)) 

        if len(exceptedFtmList) > 0:
            raise Exception("Wrong Framework Names --> ",exceptedFtmList)

        unvalidData = []
        for mapping in mappingList:
            if mapping["SrcStandardCode"] in iFrmList and mapping["DestStandardCode"] in sFrmList:
                unvalidData.append(f"Unvalid mapping. {mapping["SrcStandardCode"]} --> {mapping["DestStandardCode"]}")

        if len(unvalidData) > 0:
            raise Exception(unvalidData)

    except Exception as e:
        raise e


def PrepareSql(mappingList , sFrmList, iFrmList):
    sqlList = []
    for mapping in mappingList:   
        sSC =  mapping["SrcStandardCode"]    
        dSC = mapping["DestStandardCode"]
        sIC = mapping["SrcItemCode"]
        dIC = mapping["DestItemCode"]
        bM = mapping["BertMean"]
        sqlList.append(f"INSERT INTO TblCompStandardItemMapping([SrcStandardCode],[SrcItemCode],[DestStandardCode],[DestItemCode],[Relevence],[Confidence],[AreaScore],[ItemScore],[InsertDate],[Status])VALUES('{sSC}', '{sIC}', '{dSC}', '{dIC}', {bM} , {bM}, {bM}, {bM}, GETUTCDATE(), 1);")
        
        if not  sSC in iFrmList  and dSC in sFrmList :
            sqlList.append(f"INSERT INTO TblCompStandardItemMapping([SrcStandardCode],[SrcItemCode],[DestStandardCode],[DestItemCode],[Relevence],[Confidence],[AreaScore],[ItemScore],[InsertDate],[Status])VALUES('{dSC}', '{dIC}', '{sSC}', '{sIC}', {bM} , {bM}, {bM}, {bM}, GETUTCDATE(), 1);")
        
        sqlList.append("--" * 100)
    return sqlList
