import io
from openpyxl import load_workbook
from collections import Counter


async def ItemSqlPrepear(file, id, new_standard_name):
    try:
        contents = await file.read()
        workbook = load_workbook(io.BytesIO(contents), read_only=True)
        sheet = workbook.active

        data = []
        itemCodeList = []
        itemDescList = []
        for i, row in enumerate(sheet.iter_rows()):
            if (i == 0):
                continue
            if (row[0].value == None):
                break
            itemId = id
            sCode = new_standard_name.strip()
            iCode = row[2].value.strip()
            area = row[1].value.strip().replace("\n", "")
            desc = row[3].value.strip().replace("\n", "")

            itemCodeList.append(iCode)
            itemDescList.append(desc)
            data.append(
                f"INSERT INTO [dbo].[TblCompStandardItems]([CompItemID],[StandardCode],[ItemCode],[AreaName],[CsiosID],[InsertDate],[Description]) VALUES({itemId}, '{sCode}', '{iCode}', '{area}',1, GETUTCDATE(), '{desc}');")
            id += 1

        CheckDobleItemCode(itemCodeList)

        CheckDobleItemDescriptin(itemDescList)

        return data

    except Exception as e:
        raise Exception("error", str(e))


def CheckDobleItemCode(itemCodeList): 
    counter = Counter(itemCodeList)
    doubleItems = [t for t, c in counter.items() if c > 1]

    if len(doubleItems) > 0:
        print("Double Items:", doubleItems)
        raise Exception("There are multiple item codes..", doubleItems[0])
    
def CheckDobleItemDescriptin(itemDescList): # aynı description olan kayıtlar olabilir
    counter = Counter(itemDescList)
    doubleItems = [t for t, c in counter.items() if c > 1]

    if len(doubleItems) > 0:
        print("Double Items:", doubleItems)
        raise Exception("There are multiple item descriptions..", doubleItems[0])
